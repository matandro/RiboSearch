import numpy as np
import matplotlib.pyplot as plt
import openpyxl as xl
import os


def autolabel(rects):
    """
    Attach a text label above each bar displaying its height
    """
    for rect in rects:
        height = rect.get_height()
        ax.text(rect.get_x() + rect.get_width()/2., 1.05*height,
                '%d' % int(height),
                ha='center', va='bottom')


BASE_DIR = "D:\Matan\Dropbox\PHd\Final Thesis\data"

# Mean design score graph
wb = xl.load_workbook(os.path.join(BASE_DIR, "Seed_scores.xlsx"), read_only=True)
ws = wb['Scores']
score_lists = [[], [], []]
for row in ws.iter_rows(min_row=2, max_row=401, min_col=2, max_col=4):
    for index, cell in zip(range(0, 3), row):
        score_lists[index].append(cell.value)
score_map = {}
for index, column in zip(range(0, 3), ['B', 'C', 'D']):
    score_map[ws['{}1'.format(column)].value] = np.array(score_lists[index])
wb.close()


values = score_map.values()
CTEs = list(map(np.mean, values))
error = list(map(np.std, values))
headers = score_map.keys()
fig, ax = plt.subplots()
x_pos = np.arange(len(headers))
rects = ax.bar(x_pos, CTEs, 0.7,  color='b') # , align='center'ecolor='black',capsize=10, yerr=error, alpha=0.5,)
ax.set_ylabel('Mean design score')
ax.set_xticks(x_pos)
ax.set_ylim([0, 100])
ax.set_xticklabels(headers)
ax.set_title('Mean design scores per seed type')
#ax.yaxis.grid(True)

autolabel(rects)
# Save the figure and show
plt.tight_layout()
plt.savefig(os.path.join(BASE_DIR, 'design_score_bar.eps'), dpi=300)
plt.show()

# Mean nucleic acid change graph
seed_lists = [[], [], []]
design_lists = [[], [], []]
with open(os.path.join(BASE_DIR, "diffs_all.txt"), 'r') as input_file:
    input_file.readline()
    for line in input_file:
        words = line.strip().split('\t')
        seed_lists[0].append(int(words[0]))
        design_lists[0].append(int(words[1]))
        seed_lists[1].append(int(words[2]))
        design_lists[1].append(int(words[3]))
        seed_lists[2].append(int(words[4]))
        design_lists[2].append(int(words[5]))
seed_lists = [[item / len(single_list) for item in single_list] for single_list in seed_lists]
design_lists = [[item / len(single_list) for item in single_list] for single_list in design_lists]
seed_arrays = [np.array(single_list) for single_list in seed_lists]
design_arrays = [np.array(single_list) for single_list in design_lists]
seed_CETs = list(map(np.mean, seed_arrays))
design_CETs = list(map(np.mean, design_arrays))
seed_errors = list(map(np.std, seed_arrays))
design_errors = list(map(np.std, design_arrays))
headers = ['xpt', 'random', 'incaRNAtion']
width = 0.35

fix, ax = plt.subplots()
x_pos = np.arange(len(headers))
rects1 = ax.bar(x_pos, seed_CETs, width, color='b', yerr=seed_errors)
rects2 = ax.bar(x_pos + width, design_CETs, width, color='r', yerr=design_errors)

ax.set_ylabel('Mean nucleic acid mismatch')
ax.set_title('Mean nucleic acid mismatch per seed type')
ax.set_xticks(x_pos + width / 2)
ax.set_xticklabels(headers)
ax.legend((rects1[0], rects2[0]), ('Seed', 'Design'))

autolabel(rects1)
autolabel(rects2)

plt.tight_layout()
plt.savefig(os.path.join(BASE_DIR, 'design_diff.eps'), dpi=300)
plt.show()